import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import ttk, simpledialog
from openpyxl.utils import range_boundaries, get_column_letter
import re
import time
import os

# --- Definiciones de columnas (solo una vez, al inicio) ---
COLUMNAS_FALLOS = {
    "Fecha": 1,
    "Objeto": 2,
    "Intentos": 3,
    "Kamas Iniciales": 4,
    "Kamas Finales": 5,
    "Inversion": 6,
    "Tiempo Medio por Intento (s)": 7,
    "Kamas por Intento": 8,
    "Resultado": 9
}

COLUMNAS_EXITOS = {
    "Fecha": 1,
    "Objeto": 2,
    "Intentos Totales": 3,
    "Kamas Iniciales (acum.)": 4,
    "Kamas Finales": 5,
    "Inversion Total": 6,
    "Tiempo Medio por Intento (s)": 7,
    "Kamas por Intento (promedio)": 8,
    "Resultado": 9,
    "Precio Objeto Base": 10,
    "Precio Venta Objeto Final": 11,
    "Tipo Exo": 12,
    "Rentabilidad (kamas/h)": 13,
    "Eficiencia": 14
}


def _is_success_flag(exito):
    """Normaliza y decide si 'exito' representa un Success (incluye PA/1/True)."""
    if isinstance(exito, bool):
        return exito
    try:
        if isinstance(exito, (int, float)) and int(exito) == 1:
            return True
    except Exception:
        pass
    s = str(exito).strip().lower()
    return s in ("success", "exito", "pa", "1", "true", "ok", "exo_pa", "pa_exito")


def _safe_to_int(val, default=0):
    """Convierte val a int lo mejor posible (strings con separadores, floats, listas, dicts)."""
    if val is None:
        return default
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    # listas/tuplas: tomar primer elemento
    if isinstance(val, (list, tuple)) and len(val) > 0:
        return _safe_to_int(val[0], default)
    # dict: intentar keys comunes
    if isinstance(val, dict):
        for k in ("intentos", "attempts", "value", "kamas"):
            if k in val:
                return _safe_to_int(val[k], default)
        return default
    s = str(val).strip()
    if s == "":
        return default
    # quitar todo lo que no sea dígito o signo
    cleaned = re.sub(r'[^\d-]', '', s)
    if cleaned == "":
        return default
    try:
        return int(cleaned)
    except Exception:
        return default


def _safe_to_float(val, default=0.0):
    """Convierte val a float intentando manejar coma decimal y separadores de miles."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, (list, tuple)) and len(val) > 0:
        return _safe_to_float(val[0], default)
    if isinstance(val, dict):
        for k in ("time", "tiempo", "tiempo_medio", "value"):
            if k in val:
                return _safe_to_float(val[k], default)
        return default
    s = str(val).strip()
    if s == "":
        return default
    # eliminar separadores de miles, aceptar coma decimal
    s2 = s.replace('.', '').replace(' ', '').replace(',', '.')
    try:
        return float(s2)
    except Exception:
        return default


def agregar_datos(objeto_seleccionado, intentos, kamas_iniciales, kamas_finales, exito,
                  tiempo_medio_intento=None, modo_encadenado_activo=False,
                  precio_objeto_base=None, precio_venta_objeto_final=None, tipo_exo=None):
    """
    Agrega datos de forjam a a un archivo Excel.
    NUEVO: Acepta kamas_iniciales=None para guardados intermedios (cada 10 intentos).

    Args:
        objeto_seleccionado (str): Nombre del objeto que se está forjamagueando.
        intentos (int): Número de intentos realizados en esta sesión.
        kamas_iniciales (int): Cantidad de kamas al inicio de la sesión.
        kamas_finales (int): Cantidad de kamas al final de la sesión.
        exito (str): Resultado de la sesión ("Success" o "Fail").
        tiempo_medio_intento (float, optional): Tiempo medio por intento en segundos. Por defecto, None.
        modo_encadenado_activo (bool): True si se están realizando múltiples exos encadenados (2 o más). Por defecto, False.
        precio_objeto_base (int, optional): Precio del objeto cuando fue comprado/crafteado. Requiere para "Success" si no es modo encadenado.
        precio_venta_objeto_final (int, optional): Precio de venta del objeto exomagueado. Requiere para "Success" si no es modo encadenado.
        tipo_exo (str, optional): Tipo de exo ("PA", "PM", "AL", "NINGUNO"). Requiere para "Success" si no es modo encadenado.
    """

    # --- Configuración ---
    ruta_archivo_excel = r"C:\Users\danis\OneDrive\Desktop\Forjamagia\20260213_Dofus3_Mage_Database.xlsx"

    # CAMBIO: los fallos van a "Fallos Forjamagia" (no "Old")
    nombre_hoja_fallos = "Fallos Forjamagia"
    nombre_hoja_exitos = "Exitos Forjamagia"

    print(f"\n--- Inicia agregar_datos para '{objeto_seleccionado}', Exito: '{exito}' ---")
    print("DEBUG_RAW_INPUT TYPES/VALUES:")
    print(f"  - intentos   : type={type(intentos)} value={intentos}")
    print(f"  - kamas_init : type={type(kamas_iniciales)} value={kamas_iniciales}")
    print(f"  - kamas_fin  : type={type(kamas_finales)} value={kamas_finales}")
    print(f"  - tiempo     : type={type(tiempo_medio_intento)} value={tiempo_medio_intento}")
    print(f"  - precio_base: type={type(precio_objeto_base)} value={precio_objeto_base}")
    print(f"  - precio_vent: type={type(precio_venta_objeto_final)} value={precio_venta_objeto_final}")
    print(f"  - tipo_exo   : type={type(tipo_exo)} value={tipo_exo}")

    # Convertir de forma segura a ints/floats (tolerante a tipos inesperados)
    intentos_safe = _safe_to_int(intentos, default=0)

    # CAMBIO: Si kamas_iniciales es None, buscar el último valor guardado de kamas finales para este objeto
    if kamas_iniciales is None:
        print(f"DEBUG_KAMAS: kamas_iniciales es None. Buscando último valor guardado para '{objeto_seleccionado}'...")
        kamas_iniciales_safe = _obtener_ultimas_kamas_finales(objeto_seleccionado, ruta_archivo_excel, nombre_hoja_fallos)
        if kamas_iniciales_safe == 0:
            print("WARNING: No se encontró valor previo. Usando kamas_finales como referencia.")
            kamas_iniciales_safe = _safe_to_int(kamas_finales, default=0)
    else:
        kamas_iniciales_safe = _safe_to_int(kamas_iniciales, default=0)

    kamas_finales_safe = _safe_to_int(kamas_finales, default=0)
    tiempo_medio_safe = _safe_to_float(tiempo_medio_intento, default=0.0)
    precio_objeto_base_safe = _safe_to_int(precio_objeto_base, default=0)
    precio_venta_objeto_final_safe = _safe_to_int(precio_venta_objeto_final, default=0)

    print("DEBUG_COERCED VALUES:")
    print(f"  - intentos   : {intentos_safe}")
    print(f"  - kamas_init : {kamas_iniciales_safe}")
    print(f"  - kamas_fin  : {kamas_finales_safe}")
    print(f"  - tiempo     : {tiempo_medio_safe}")
    print(f"  - precio_base: {precio_objeto_base_safe}")
    print(f"  - precio_vent: {precio_venta_objeto_final_safe}")

    # Reassign cleaned values for rest of function
    try:
        intentos = int(intentos_safe)
        kamas_iniciales = int(kamas_iniciales_safe)
        kamas_finales = int(kamas_finales_safe)
    except Exception as e:
        print(f"ERROR: No se pudieron convertir valores a entero: {e}")
        return False

    tiempo_medio_intento = float(tiempo_medio_safe)
    precio_objeto_base = int(precio_objeto_base_safe)
    precio_venta_objeto_final = int(precio_venta_objeto_final_safe)

    # NUEVO: bandera de seguridad para lecturas inválidas/interrumpidas
    kamas_finales_en_cero = (kamas_finales == 0)

    # --- Cargar o crear el archivo y las hojas de Excel ---
    # Intentar abrir el archivo; si está bloqueado por Excel, esperar hasta que se cierre.
    libro_excel = None
    while True:
        try:
            if os.path.exists(ruta_archivo_excel):
                libro_excel = openpyxl.load_workbook(ruta_archivo_excel)
                print(f"DEBUG_EXCEL: Archivo '{ruta_archivo_excel}' cargado exitosamente.")
            else:
                print(f"Advertencia: El archivo '{ruta_archivo_excel}' no se encontró. Creando uno nuevo.")
                libro_excel = openpyxl.Workbook()
                if "Sheet" in libro_excel.sheetnames:
                    libro_excel.remove(libro_excel["Sheet"])
                print(f"DEBUG_EXCEL: Archivo '{ruta_archivo_excel}' creado (en memoria).")
            break
        except (PermissionError, OSError) as e:
            # Archivo probablemente abierto en Excel; esperar y reintentar
            print(f"ERROR_OPEN: No se puede abrir '{ruta_archivo_excel}' (posible bloqueo): {e}")
            print("DEBUG_OPEN: Esperando a que se cierre el archivo para poder abrirlo. Pulsa Ctrl+C para cancelar.")
            try:
                time.sleep(2)
                continue
            except KeyboardInterrupt:
                print("Operación cancelada por usuario durante espera de apertura de archivo.")
                return False
        except Exception as e:
            # Error inesperado: intentar crear un libro nuevo en memoria
            print(f"ERROR_OPEN: Error inesperado al abrir '{ruta_archivo_excel}': {e}")
            try:
                libro_excel = openpyxl.Workbook()
                if "Sheet" in libro_excel.sheetnames:
                    libro_excel.remove(libro_excel["Sheet"])
                print("DEBUG_EXCEL: Se creó un libro nuevo en memoria tras error de apertura.")
                break
            except Exception as e2:
                print(f"ERROR_OPEN: No se pudo crear un libro nuevo en memoria: {e2}")
                return False

    # Ensure "Fallos Forjamagia" sheet exists and has headers
    if nombre_hoja_fallos not in libro_excel.sheetnames:
        hoja_fallos = libro_excel.create_sheet(nombre_hoja_fallos)
        for col_name, col_idx in COLUMNAS_FALLOS.items():
            hoja_fallos.cell(row=1, column=col_idx, value=col_name)
        print(f"DEBUG_EXCEL: Hoja '{nombre_hoja_fallos}' creada y cabeceras añadidas.")
    else:
        hoja_fallos = libro_excel[nombre_hoja_fallos]
        if not hoja_fallos.cell(1, COLUMNAS_FALLOS["Fecha"]).value:  # Check if header row is empty
            for col_name, col_idx in COLUMNAS_FALLOS.items():
                hoja_fallos.cell(row=1, column=col_idx, value=col_name)
            print(f"DEBUG_EXCEL: Cabeceras añadidas a la hoja existente '{nombre_hoja_fallos}'.")
        else:
            print(f"DEBUG_EXCEL: Hoja '{nombre_hoja_fallos}' ya existe con cabeceras.")

    # Ensure "Exitos Forjamagia" sheet exists and has headers
    if nombre_hoja_exitos not in libro_excel.sheetnames:
        hoja_exitos = libro_excel.create_sheet(nombre_hoja_exitos)
        for col_name, col_idx in COLUMNAS_EXITOS.items():
            hoja_exitos.cell(row=1, column=col_idx, value=col_name)
        print(f"DEBUG_EXCEL: Hoja '{nombre_hoja_exitos}' creada y cabeceras añadidas.")
    else:
        hoja_exitos = libro_excel[nombre_hoja_exitos]
        if not hoja_exitos.cell(1, COLUMNAS_EXITOS["Fecha"]).value:  # Check if header row is empty
            for col_name, col_idx in COLUMNAS_EXITOS.items():
                hoja_exitos.cell(row=1, column=col_idx, value=col_name)
            print(f"DEBUG_EXCEL: Cabeceras añadidas a la hoja existente '{nombre_hoja_exitos}'.")
        else:
            print(f"DEBUG_EXCEL: Hoja '{nombre_hoja_exitos}' ya existe con cabeceras.")

    # --- Variables comunes para la sesión actual ---
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")

    # Inversión de la fila actual (fallo o éxito del ciclo actual)
    if kamas_finales_en_cero:
        inversion_actual_sesion = 0
        kamas_por_intento_actual_sesion = 0.0
        print("DEBUG_GUARD: kamas_finales=0 detectado -> inversion_actual_sesion=0, kamas_por_intento_actual_sesion=0")
    else:
        inversion_actual_sesion = kamas_iniciales - kamas_finales
        kamas_por_intento_actual_sesion = inversion_actual_sesion / intentos if intentos > 0 else 0

    tiempo_por_intento_actual_sesion = tiempo_medio_intento if tiempo_medio_intento is not None else 0.0  # Ensure float

    print(f"DEBUG_CALC: Inversion de la sesión actual: {inversion_actual_sesion}")
    print(f"DEBUG_CALC: Kamas por intento de la sesión actual: {kamas_por_intento_actual_sesion:.2f}")
    print(f"DEBUG_CALC: Tiempo por intento de la sesión actual: {tiempo_por_intento_actual_sesion:.2f}")

    # --- Lógica de inserción/actualización ---
    es_success = _is_success_flag(exito)

    if not es_success:
        print(f"DEBUG_FLOW: Resultado considerado 'Fail'. Añadiendo a la hoja '{nombre_hoja_fallos}'.")

        # CAMBIO: Usar kamas_iniciales_safe (ya corregido arriba) en lugar de fallback a kamas_finales
        kamas_iniciales_display = kamas_iniciales_safe if kamas_iniciales_safe > 0 else kamas_finales

        # CAMBIO: inversión en fallos siempre basada en la fila (no en acumulado global)
        if kamas_finales_en_cero:
            inversion_fila = 0
            kamas_por_intento_fila = 0.0
            print("DEBUG_GUARD: FAIL con kamas_finales=0 -> inversion_fila=0, kamas_por_intento_fila=0")
        else:
            inversion_fila = kamas_iniciales_display - kamas_finales
            kamas_por_intento_fila = inversion_fila / intentos if intentos > 0 else 0.0

        default_row_values = {
            "Fecha": fecha_hoy,
            "Objeto": objeto_seleccionado,
            "Intentos": intentos,
            "Kamas Iniciales": kamas_iniciales_display,
            "Kamas Finales": kamas_finales,
            "Inversion": inversion_fila,
            "Tiempo Medio por Intento (s)": tiempo_por_intento_actual_sesion or 0.0,
            "Kamas por Intento": kamas_por_intento_fila,
            "Resultado": str(exito)
        }

        expected_headers_fallos = list(COLUMNAS_FALLOS.keys())
        table, min_c, min_r, max_c, max_r = _find_table_for_headers(hoja_fallos, expected_headers_fallos)

        if table is not None:
            try:
                _insert_row_into_table(hoja_fallos, table, min_c, min_r, max_c, max_r, default_row_values)
                print(f"DEBUG_FLOW: Registro de fallo insertado dentro de la tabla en '{nombre_hoja_fallos}'.")
            except Exception as e:
                print("WARNING: insertar en tabla falló, usando append. Error:", e)
                fila_destino = hoja_fallos.max_row + 1
                # usar header row para escribir si existe
                header_row = min_r if table is not None else 1
                _write_row_using_headers(hoja_fallos, header_row, fila_destino, default_row_values)
                print(f"DEBUG_FLOW: Registro de fallo para '{objeto_seleccionado}' agregado a '{nombre_hoja_fallos}' (append).")
        else:
            fila_destino = hoja_fallos.max_row + 1
            _write_row_by_expected_mapping(hoja_fallos, fila_destino, COLUMNAS_FALLOS, default_row_values)
            print(f"DEBUG_FLOW: Registro de fallo para '{objeto_seleccionado}' agregado a '{nombre_hoja_fallos}' (sheet end).")

    else:
        print(f"DEBUG_FLOW: Resultado considerado 'Success'. Procesando para la hoja '{nombre_hoja_exitos}'.")

        # Si el exito indica PA y no se pasó tipo_exo, setear 'PA'
        if tipo_exo is None:
            try:
                s = str(exito).strip().lower()
                if "pa" in s or str(exito) == "1":
                    tipo_exo = "PA"
            except Exception:
                tipo_exo = tipo_exo or "N/A"

        # En modo encadenado no pedimos precios
        if modo_encadenado_activo:
            precio_objeto_base = precio_objeto_base if precio_objeto_base is not None else 0
            precio_venta_objeto_final = precio_venta_objeto_final if precio_venta_objeto_final is not None else 0
            tipo_exo = tipo_exo if tipo_exo is not None else "N/A (Encadenado)"
        else:
            # If values missing, fallback to 0 or "N/A" to avoid blocking; caller can fix later
            precio_objeto_base = precio_objeto_base if precio_objeto_base is not None else 0
            precio_venta_objeto_final = precio_venta_objeto_final if precio_venta_objeto_final is not None else 0
            tipo_exo = tipo_exo if tipo_exo is not None else "N/A"

        # CAMBIO: preparar acumulación desde Old con orden cronológico real
        total_intentos_acumulados = intentos
        total_tiempo_acumulado = (tiempo_por_intento_actual_sesion or 0.0) * intentos
        inversion_acumulada = inversion_actual_sesion

        # Regla pedida:
        # - kamas iniciales = primero (más antiguo en Old)
        # - kamas finales = último (éxito actual)
        primer_kamas_inicial = None
        kamas_finales_exito = kamas_finales

        rows_objeto = []
        for fila_num in range(2, hoja_fallos.max_row + 1):
            val_obj = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Objeto"]).value
            if str(val_obj).strip() == str(objeto_seleccionado).strip():
                rows_objeto.append(fila_num)

        for fila_num in rows_objeto:
            current_fail_attempts_val = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Intentos"]).value
            inversion_fail_val = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Inversion"]).value
            tiempo_medio_intento_fail_val = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Tiempo Medio por Intento (s)"]).value
            kamas_iniciales_fail_val = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Kamas Iniciales"]).value

            fail_attempts = _safe_to_int(current_fail_attempts_val, default=0)
            total_intentos_acumulados += fail_attempts

            # CAMBIO CLAVE: inversión total = SUMA de todas las inversiones de filas
            inversion_acumulada += _safe_to_float(inversion_fail_val, default=0.0)

            fail_tmedio = _safe_to_float(tiempo_medio_intento_fail_val, default=0.0)
            total_tiempo_acumulado += (fail_tmedio * fail_attempts)

            if primer_kamas_inicial is None:
                primer_kamas_inicial = _safe_to_int(kamas_iniciales_fail_val, default=0)

        if primer_kamas_inicial is None:
            primer_kamas_inicial = kamas_iniciales

        # borrar filas fusionadas (de abajo a arriba)
        for fila_num in sorted(rows_objeto, reverse=True):
            hoja_fallos.delete_rows(fila_num)
        print(f"DEBUG_FLOW: Eliminadas {len(rows_objeto)} filas fusionadas de '{nombre_hoja_fallos}' para '{objeto_seleccionado}'.")

        total_tiempo_por_intento_acumulado = (total_tiempo_acumulado / total_intentos_acumulados) if total_intentos_acumulados > 0 else 0.0
        kamas_por_intento_acumulado = (inversion_acumulada / total_intentos_acumulados) if total_intentos_acumulados > 0 else 0.0

        rentabilidad = None
        if total_tiempo_acumulado > 0:
            horas = total_tiempo_acumulado / 3600.0
            net_kamas = kamas_finales_exito - primer_kamas_inicial
            try:
                rentabilidad = net_kamas / horas
            except Exception:
                rentabilidad = None

        eficiencia = None
        try:
            if rentabilidad is not None:
                eficiencia = float(rentabilidad) / (abs(kamas_por_intento_acumulado) + 1)
        except Exception:
            eficiencia = None

        # CAMBIO: KPI base vs KPI robusto por fallos
        kamas_por_intento_base = (inversion_acumulada / total_intentos_acumulados) if total_intentos_acumulados > 0 else 0.0
        kamas_por_intento_fallos_robusto = _kpi_fallos_robusto_para_objeto(hoja_fallos, objeto_seleccionado)

        if kamas_por_intento_fallos_robusto is not None:
            kamas_por_intento_acumulado = kamas_por_intento_fallos_robusto
            print(f"DEBUG_KPI: usando media robusta fallos para KPI. base={kamas_por_intento_base:.2f} robusto={kamas_por_intento_acumulado:.2f}")
        else:
            kamas_por_intento_acumulado = kamas_por_intento_base
            print(f"DEBUG_KPI: sin histórico robusto de fallos. usando KPI base={kamas_por_intento_acumulado:.2f}")

        # NUEVO: regla pedida para éxito con kamas_finales=0
        if kamas_finales_en_cero:
            inversion_acumulada = 0.0
            kamas_por_intento_acumulado = 0.0
            print("DEBUG_GUARD: SUCCESS con kamas_finales=0 -> Inversion Total=0, Kamas por Intento (promedio)=0")

        row_values_exito = {
            "Fecha": fecha_hoy,
            "Objeto": objeto_seleccionado,
            "Intentos Totales": total_intentos_acumulados,
            "Kamas Iniciales (acum.)": primer_kamas_inicial,
            "Kamas Finales": kamas_finales_exito,
            "Inversion Total": inversion_acumulada,
            "Tiempo Medio por Intento (s)": total_tiempo_por_intento_acumulado,
            "Kamas por Intento (promedio)": kamas_por_intento_acumulado,  # <- robustecido / protegido
            "Resultado": str(exito),
            "Precio Objeto Base": precio_objeto_base if precio_objeto_base is not None else 0,
            "Precio Venta Objeto Final": precio_venta_objeto_final if precio_venta_objeto_final is not None else 0,
            "Tipo Exo": tipo_exo,
            "Rentabilidad (kamas/h)": rentabilidad,
            "Eficiencia": eficiencia
        }

        expected_headers_exitos = list(COLUMNAS_EXITOS.keys())
        table_e, min_ce, min_re, max_ce, max_re = _find_table_for_headers(hoja_exitos, expected_headers_exitos)

        if table_e is not None:
            try:
                _insert_row_into_table(hoja_exitos, table_e, min_ce, min_re, max_ce, max_re, row_values_exito)
                print(f"DEBUG_FLOW: Registro de success insertado dentro de la tabla en '{nombre_hoja_exitos}'.")
            except Exception as e:
                print("WARNING: insertar en tabla éxitos falló, usando append. Error:", e)
                fila_destino = hoja_exitos.max_row + 1
                # si la hoja tiene cabeceras, escribir por nombre usando header row
                header_row = min_re if table_e is not None else 1
                _write_row_using_headers(hoja_exitos, header_row, fila_destino, row_values_exito)
                print(f"DEBUG_FLOW: Registro de success para '{objeto_seleccionado}' agregado a '{nombre_hoja_exitos}' (append).")
        else:
            fila_destino = hoja_exitos.max_row + 1
            # usar mapping esperado (COLUMNAS_EXITOS) si la hoja no tiene tabla identificable
            _write_row_by_expected_mapping(hoja_exitos, fila_destino, COLUMNAS_EXITOS, row_values_exito)
            print(f"DEBUG_FLOW: Registro de success para '{objeto_seleccionado}' agregado a '{nombre_hoja_exitos}' (sheet end).")

    # --- Guardar el archivo Excel con espera si está abierto ---
    # Intentar guardar; si el archivo está abierto en Excel (PermissionError/OSError), esperar hasta que se cierre.
    print(f"DEBUG_SAVE: Intentando guardar datos en '{ruta_archivo_excel}'...")
    save_success = False
    while True:
        try:
            libro_excel.save(ruta_archivo_excel)
            print(f"DEBUG_SAVE: Datos guardados exitosamente en '{ruta_archivo_excel}'.")
            save_success = True
            break
        except PermissionError as e:
            print(f"ERROR_SAVE: Archivo probablemente abierto en Excel: {e}")
            print("DEBUG_SAVE: Esperando a que se cierre el archivo para poder guardar. Pulsa Ctrl+C para cancelar.")
            try:
                time.sleep(2)
            except KeyboardInterrupt:
                print("Operación cancelada por usuario durante espera de cierre de archivo.")
                break
        except OSError as e:
            # Algunos errores de Windows salen como OSError; tratarlos igual que PermissionError
            print(f"ERROR_SAVE: OSError al guardar (posiblemente archivo bloqueado): {e}")
            print("DEBUG_SAVE: Esperando a que se cierre el archivo para poder guardar. Pulsa Ctrl+C para cancelar.")
            try:
                time.sleep(2)
            except KeyboardInterrupt:
                print("Operación cancelada por usuario durante espera de cierre de archivo.")
                break
        except Exception as e:
            # Si hay otro error inesperado, intentar guardar con fallback timestamp como último recurso
            print(f"ERROR_SAVE: Error inesperado al guardar: {e}")
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fallback_path = ruta_archivo_excel.replace(".xlsx", f".{ts}.xlsx")
                libro_excel.save(fallback_path)
                print(f"DEBUG_SAVE: Fallback guardado en '{fallback_path}'.")
                save_success = True
                break
            except Exception as e2:
                print(f"ERROR_SAVE: Fallback también falló: {e2}")
                print("ERROR_SAVE: Asegúrate de que el archivo no esté abierto en otra aplicación (como Excel).")
                print(f"--- Finaliza agregar_datos para '{objeto_seleccionado}' con ERROR ---\n")
                save_success = False
                break

    if save_success:
        print(f"--- Finaliza agregar_datos para '{objeto_seleccionado}' ---\n")
        return True
    else:
        return False


# -------------------------------------------------------------------------
# Nuevos helpers robustos: colocarlos justo después de los imports
# -------------------------------------------------------------------------
def _normalize_value_for_excel(value):
    """
    Convierte strings numéricos con separadores (., espacios, ,) a int/float.
    Devuelve '' para None. Mantiene tipos numéricos.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if s == "":
        return ""
    # eliminar separadores de miles y espacios
    cleaned = re.sub(r'[.\s]', '', s)
    # si quedan solo dígitos -> int
    if cleaned.isdigit():
        try:
            return int(cleaned)
        except Exception:
            pass
    # intentar float (coma decimal -> punto)
    alt = s.replace(',', '.')
    try:
        f = float(alt)
        return f
    except Exception:
        pass
    # fallback: devolver string limpio (sin puntos de miles)
    return re.sub(r'\s+', ' ', s)

def _find_table_for_headers(sheet, expected_headers):
    """
    Busca una tabla cuyo encabezado contenga (en orden flexible) expected_headers.
    Devuelve (table, min_col, min_row, max_col, max_row) o (None, None, None, None, None).
    """
    expected_norm = [h.strip().lower() for h in expected_headers]
    for table in getattr(sheet, "_tables", []):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except Exception:
            continue
        header_cells = [sheet.cell(row=min_row, column=col).value for col in range(min_col, max_col + 1)]
        header_norm = [str(h).strip().lower() if h else "" for h in header_cells]
        # comprobar que todos los expected headers aparecen in orden (flexible)
        idx = 0
        ok = True
        for exp in expected_norm:
            found = False
            while idx < len(header_norm):
                if exp == header_norm[idx] or exp in header_norm[idx]:
                    found = True
                    idx += 1
                    break
                idx += 1
            if not found:
                ok = False
                break
        if ok:
            return table, min_col, min_row, max_col, max_row
    return None, None, None, None, None

def _insert_row_into_table(sheet, table, min_col, min_row, max_col, max_row, row_values_by_header):
    """
    Inserta una fila dentro de la tabla y actualiza table.ref.
    Escribe usando los textos de cabecera presentes en la tabla (min_row).
    """
    insert_at = max_row + 1
    sheet.insert_rows(insert_at)
    for col in range(min_col, max_col + 1):
        header = sheet.cell(row=min_row, column=col).value
        key = str(header).strip() if header is not None else ""
        raw_value = row_values_by_header.get(key, "")
        value = _normalize_value_for_excel(raw_value)
        sheet.cell(row=insert_at, column=col, value=value)
    # actualizar referencia de la tabla
    new_end_row = max_row + 1
    try:
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_end_row}"
        table.ref = new_ref
    except Exception as e:
        print("WARNING: no se pudo actualizar table.ref:", e)

def _write_row_using_headers(sheet, header_row, row_idx, row_values_by_header):
    """
    Escribe una fila usando los nombres de la fila de cabecera (header_row).
    Si la hoja tiene más columnas que las keys, rellena con ''.
    """
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        header = sheet.cell(row=header_row, column=col).value
        if header is None:
            continue
        key = str(header).strip()
        raw = row_values_by_header.get(key, "")
        sheet.cell(row=row_idx, column=col, value=_normalize_value_for_excel(raw))

def _write_row_by_expected_mapping(sheet, row_idx, mapping_dict, row_values):
    """
    Escribe una fila usando mapping_dict (header_name -> col_idx).
    Convierte valores con _normalize_value_for_excel.
    """
    for header_name, col_idx in mapping_dict.items():
        raw = row_values.get(header_name, "")
        sheet.cell(row=row_idx, column=col_idx, value=_normalize_value_for_excel(raw))


def run_two_phase_demo(objeto="DemoObjeto_TwoPhase", n_fail=10):
    """
    Prueba en dos partes:
      1) Inserta n_fail registros de fallo para 'objeto'.
      2) Pregunta al usuario y, si confirma, inserta un registro de éxito PA que
         acumulará y eliminará los fallos previos del mismo objeto.
    """
    print(f"\n--- RUN TWO PHASE DEMO: objeto='{objeto}', fallos={n_fail} ---")
    print("Fase 1: añadiendo fallos...")
    for i in range(1, n_fail + 1):
        # Valores de ejemplo; ajusta si lo deseas
        intentos = 1
        kamas_iniciales = 1_000_000
        kamas_finales = kamas_iniciales - (i * 100)  # pequeño decremento por fallo
        tiempo_med = 2.0
        ok = agregar_datos(
            objeto_seleccionado=objeto,
            intentos=intentos,
            kamas_iniciales=kamas_iniciales,
            kamas_finales=kamas_finales,
            exito=0,  # fallo
            tiempo_medio_intento=tiempo_med,
            modo_encadenado_activo=True
        )
        print(f"  -> Fallo {i}/{n_fail} añadido (guardado={bool(ok)})")
    print("Fase 1 completada. Ahora hay múltiples filas de fallos para el mismo objeto.")

    entrada = input("Escribe 'ok' para proceder con la Fase 2 (inserción de EXITO PA y acumulación), o Enter para cancelar: ").strip().lower()
    if entrada != "ok":
        print("Operación cancelada por el usuario. No se realizó la fase 2.")
        return

    print("Fase 2: insertando registro de ÉXITO PA y acumulando fallos previos...")
    # Valores de ejemplo para el éxito; modo_encadenado evita popups
    result_ok = agregar_datos(
        objeto_seleccionado=objeto,
        intentos=1,
        kamas_iniciales=1_000_000,
        kamas_finales=1_000_000 - (n_fail * 100) - 500,  # ejemplo: menor que todos los previos
        exito="PA",
        tiempo_medio_intento=2.0,
        modo_encadenado_activo=True,
        precio_objeto_base=50000,
        precio_venta_objeto_final=150000,
        tipo_exo="PA"
    )
    if result_ok:
        print("Fase 2 completada: registro de éxito insertado y fallos previos acumulados/eliminados.")
    else:
        print("Fase 2 falló: revisar mensajes de DEBUG_SAVE/ERROR_SAVE en la consola.")

# Reemplazo del bloque __main__ previo por la prueba en dos fases
if __name__ == "__main__":
    print("Ejecutando prueba interactiva de Extra_Database: inserta fallos y luego éxito que acumula.")
    nombre_obj = input("Nombre del objeto para la prueba (por defecto 'DemoObjeto_TwoPhase'): ").strip() or "DemoObjeto_TwoPhase"
    try:
        n = int(input("Cuántos fallos añadir en fase 1? [10]: ").strip() or "10")
    except ValueError:
        n = 10
    run_two_phase_demo(objeto=nombre_obj, n_fail=n)

    # --- PARTE 2: Migración de datos de "Old" a "Exitos" ---
    print("\n" + "="*60)
    print("PARTE 2: Migración de datos de hoja 'Old' a 'Exitos Forjamagia'")
    print("="*60)
    input("Pulsa Enter para ejecutar la migración de 'Old' -> 'Exitos' (Ctrl+C para cancelar)...")

    ruta_archivo_excel = r"C:\Users\danis\OneDrive\Desktop\Forjamagia\20260213_Dofus3_Mage_Database.xlsx"
    nombre_hoja_old = "Old"
    nombre_hoja_exitos = "Exitos Forjamagia"

    try:
        if not os.path.exists(ruta_archivo_excel):
            print(f"ERROR: Archivo '{ruta_archivo_excel}' no encontrado. Abortando migración.")
        else:
            libro = openpyxl.load_workbook(ruta_archivo_excel)
            if nombre_hoja_old not in libro.sheetnames:
                print(f"ERROR: Hoja '{nombre_hoja_old}' no encontrada en el archivo. Abortando.")
            else:
                hoja_old = libro[nombre_hoja_old]
                acumulados = {}

                for row_num in range(2, hoja_old.max_row + 1):
                    obj_val = hoja_old.cell(row=row_num, column=2).value
                    if not obj_val:
                        continue
                    obj_name = str(obj_val).strip()
                    if obj_name not in acumulados:
                        acumulados[obj_name] = {
                            "intentos": 0,
                            "kamas_iniciales_min": 10**9,
                            "kamas_finales": 0,
                            "inversion_total": 0,
                            "tiempo_total": 0.0,
                            "rows": []
                        }

                    acc = acumulados[obj_name]
                    acc["rows"].append(row_num)

                    try:
                        intentos_val = _safe_to_int(hoja_old.cell(row=row_num, column=3).value, default=0)
                        acc["intentos"] += intentos_val
                    except Exception:
                        pass

                    try:
                        kamas_init_val = _safe_to_int(hoja_old.cell(row=row_num, column=4).value, default=0)
                        acc["kamas_iniciales_min"] = min(acc["kamas_iniciales_min"], kamas_init_val)
                    except Exception:
                        pass

                    try:
                        kamas_fin_val = _safe_to_int(hoja_old.cell(row=row_num, column=5).value, default=0)
                        acc["kamas_finales"] = kamas_fin_val
                    except Exception:
                        pass

                    try:
                        inversion_val = _safe_to_int(hoja_old.cell(row=row_num, column=6).value, default=0)
                        acc["inversion_total"] += inversion_val
                    except Exception:
                        pass

                    try:
                        tiempo_val = _safe_to_float(hoja_old.cell(row=row_num, column=7).value, default=0.0)
                        acc["tiempo_total"] += tiempo_val * intentos_val
                    except Exception:
                        pass

                print(f"\nEncontrados {len(acumulados)} objetos únicos en 'Old'.")

                if nombre_hoja_exitos not in libro.sheetnames:
                    hoja_exitos = libro.create_sheet(nombre_hoja_exitos)
                    for col_name, col_idx in COLUMNAS_EXITOS.items():
                        hoja_exitos.cell(row=1, column=col_idx, value=col_name)
                else:
                    hoja_exitos = libro[nombre_hoja_exitos]

                fecha_migracion = datetime.now().strftime("%Y-%m-%d")

                for obj_name, acc in acumulados.items():
                    total_intentos = acc["intentos"]
                    kamas_iniciales_acum = acc["kamas_iniciales_min"] if acc["kamas_iniciales_min"] < 10**9 else 0
                    kamas_finales = acc["kamas_finales"]
                    inversion_total = acc["inversion_total"]
                    tiempo_total = acc["tiempo_total"]

                    # Calcular promedios
                    tiempo_medio_intento = (tiempo_total / total_intentos) if total_intentos > 0 else 0.0
                    kamas_por_intento = (inversion_total / total_intentos) if total_intentos > 0 else 0.0

                    # Calcular rentabilidad
                    rentabilidad = None
                    if tiempo_total > 0:
                        horas = tiempo_total / 3600.0
                        net_kamas = kamas_finales - kamas_iniciales_acum
                        try:
                            rentabilidad = net_kamas / horas
                        except Exception:
                            rentabilidad = None

                    # Calcular eficiencia
                    eficiencia = None
                    try:
                        if rentabilidad is not None:
                            eficiencia = float(rentabilidad) / (abs(kamas_por_intento) + 1)
                    except Exception:
                        eficiencia = None

                    # Construir row_values (sin precios/tipo_exo ya que vienen de Old)
                    row_values_exito = {
                        "Fecha": fecha_migracion,
                        "Objeto": obj_name,
                        "Intentos Totales": total_intentos,
                        "Kamas Iniciales (acum.)": kamas_iniciales_acum,
                        "Kamas Finales": kamas_finales,
                        "Inversion Total": inversion_total,
                        "Tiempo Medio por Intento (s)": tiempo_medio_intento,
                        "Kamas por Intento (promedio)": kamas_por_intento,
                        "Resultado": "PA (migrado)",
                        "Precio Objeto Base": 0,
                        "Precio Venta Objeto Final": 0,
                        "Tipo Exo": "PA",
                        "Rentabilidad (kamas/h)": rentabilidad,
                        "Eficiencia": eficiencia
                    }

                    # Insertar en hoja de éxitos (al final)
                    fila_destino = hoja_exitos.max_row + 1
                    _write_row_by_expected_mapping(hoja_exitos, fila_destino, COLUMNAS_EXITOS, row_values_exito)
                    print(f"  Migrado éxito PA para '{obj_name}': {total_intentos} intentos, rentabilidad={rentabilidad}, eficiencia={eficiencia}")

                # Eliminar filas procesadas de "Old" (de abajo hacia arriba para evitar desajustes de índices)
                all_rows_to_delete = []
                for acc in acumulados.values():
                    all_rows_to_delete.extend(acc["rows"])
                all_rows_to_delete = sorted(set(all_rows_to_delete), reverse=True)

                for row_num in all_rows_to_delete:
                    hoja_old.delete_rows(row_num)
                print(f"\nEliminadas {len(all_rows_to_delete)} filas de la hoja '{nombre_hoja_old}'.")

                # Guardar archivo
                libro.save(ruta_archivo_excel)
                print(f"\nMigración completada. Datos guardados en '{ruta_archivo_excel}'.")

    except Exception as e:
        print(f"ERROR durante la migración: {e}")

def _obtener_ultimas_kamas_finales(objeto_nombre, ruta_excel, nombre_hoja):
    """
    Busca la última fila guardada para el objeto dado y devuelve sus kamas finales.
    Retorna 0 si no encuentra nada.
    """
    try:
        if not os.path.exists(ruta_excel):
            return 0

        libro = openpyxl.load_workbook(ruta_excel, read_only=True)
        if nombre_hoja not in libro.sheetnames:
            libro.close()
            return 0

        hoja = libro[nombre_hoja]

        # Buscar de abajo hacia arriba (más reciente primero)
        for fila_num in range(hoja.max_row, 1, -1):
            obj_val = hoja.cell(row=fila_num, column=COLUMNAS_FALLOS["Objeto"]).value
            if str(obj_val).strip() == str(objeto_nombre).strip():
                kamas_fin = hoja.cell(row=fila_num, column=COLUMNAS_FALLOS["Kamas Finales"]).value
                libro.close()
                return _safe_to_int(kamas_fin, default=0)

        libro.close()
        return 0
    except Exception as e:
        print(f"WARNING: Error buscando últimas kamas finales: {e}")
        return 0

def _compute_trimmed_mean(values, trim_ratio=0.2):
    """
    Media recortada simple para reducir impacto de outliers.
    Si hay pocos valores, cae a media normal.
    """
    vals = [float(v) for v in values if v is not None]
    if not vals:
        return None
    vals.sort()
    n = len(vals)
    k = int(n * trim_ratio)
    if n - 2 * k <= 0:
        return sum(vals) / n
    core = vals[k:n - k]
    return (sum(core) / len(core)) if core else (sum(vals) / n)


def _kpi_fallos_robusto_para_objeto(hoja_fallos, objeto_seleccionado):
    """
    Devuelve media robusta de 'Kamas por Intento' para el objeto leyendo filas de fallos.
    Fallback: si la columna viene vacía, usa inversion/intentos de la fila.
    """
    kpi_vals = []
    for fila_num in range(2, hoja_fallos.max_row + 1):
        obj_val = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Objeto"]).value
        if str(obj_val).strip() != str(objeto_seleccionado).strip():
            continue

        kpi_cell = hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Kamas por Intento"]).value
        kpi = _safe_to_float(kpi_cell, default=None)
        if kpi is None:
            inv = _safe_to_float(hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Inversion"]).value, default=0.0)
            it = _safe_to_int(hoja_fallos.cell(row=fila_num, column=COLUMNAS_FALLOS["Intentos"]).value, default=0)
            kpi = (inv / it) if it > 0 else None

        if kpi is not None:
            kpi_vals.append(kpi)

    if not kpi_vals:
        return None
    return _compute_trimmed_mean(kpi_vals, trim_ratio=0.2)
